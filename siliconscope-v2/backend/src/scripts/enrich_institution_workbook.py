#!/usr/bin/env python3
"""Clean and geocode the SiliconScope institution workbook from a ROR data dump."""

from __future__ import annotations

import argparse
import csv
import html
import math
import re
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


EMPTY_VALUES = {"", "n/a", "na", "none", "null", "unknown", "-"}
GENERIC_TOKENS = {
    "a", "an", "and", "at", "center", "centre", "co", "company", "corporation",
    "department", "for", "group", "hospital", "in", "inc", "institute", "institution",
    "international", "lab", "laboratories", "laboratory", "limited", "ltd", "of", "research",
    "school", "the", "technology", "university",
}
COUNTRY_SUFFIXES = {
    "united states": "US", "usa": "US", "u s a": "US", "us": "US",
    "united kingdom": "GB", "uk": "GB", "england": "GB",
    "south korea": "KR", "korea": "KR", "republic of korea": "KR",
    "china": "CN", "mainland china": "CN", "taiwan": "TW", "hong kong": "HK",
    "macao": "MO", "macau": "MO", "singapore": "SG", "japan": "JP",
    "france": "FR", "germany": "DE", "netherlands": "NL", "belgium": "BE",
    "switzerland": "CH", "italy": "IT", "canada": "CA", "australia": "AU",
    "india": "IN", "israel": "IL", "spain": "ES", "sweden": "SE",
}
MANUAL_OVERRIDES = {
    "university of electronic science and technology of china": ("University of Electronic Science and Technology of China", "CN", "China", "Chengdu", 30.6667, 104.0667, "UESTC"),
    "chinese university of hong kong shenzhen": ("The Chinese University of Hong Kong, Shenzhen", "CN", "China", "Shenzhen", 22.5333, 114.0667, "CUHK-Shenzhen"),
    "chinese university of hong kong": ("The Chinese University of Hong Kong", "HK", "Hong Kong", "Hong Kong", 22.4196, 114.2068, "CUHK"),
    "korea advanced science and technology": ("Korea Advanced Institute of Science and Technology", "KR", "South Korea", "Daejeon", 36.3504, 127.3845, "KAIST"),
    "georgia technology": ("Georgia Institute of Technology", "US", "United States", "Atlanta", 33.749, -84.388, "Georgia Tech"),
    "imec": ("imec", "BE", "Belgium", "Leuven", 50.8798, 4.7005, "imec"),
    "virginia technology": ("Virginia Polytechnic Institute and State University", "US", "United States", "Blacksburg", 37.2296, -80.4139, "Virginia Tech"),
    "tokyo technology": ("Institute of Science Tokyo", "JP", "Japan", "Tokyo", 35.6764, 139.6500, "Science Tokyo"),
    "national de la recherche scientifique": ("French National Centre for Scientific Research", "FR", "France", "Paris", 48.8566, 2.3522, "CNRS"),
    "microelectronics": ("Institute of Microelectronics", "SG", "Singapore", "Singapore", 1.28967, 103.85007, "IME"),
    "northeastern university": ("Northeastern University", "US", "United States", "Boston", 42.35843, -71.05977, "NEU"),
    "northwestern university": ("Northwestern University", "US", "United States", "Evanston", 42.04114, -87.69006, "NU"),
    "university of connecticut": ("University of Connecticut", "US", "United States", "Storrs", 41.80843, -72.24952, "UConn"),
    "industrial technology research institute": ("Industrial Technology Research Institute", "TW", "Taiwan", "Hsinchu", 24.80361, 120.96861, "ITRI"),
    "western university": ("Western University", "CA", "Canada", "London", 42.98339, -81.23304, "Western"),
    "soochow university": ("Soochow University", "CN", "China", "Suzhou", 31.30408, 120.59538, "SUDA"),
    "nile university": ("Nile University", "EG", "Egypt", "Giza", 30.00944, 31.20861, "NU"),
    "chang an university": ("Chang'an University", "CN", "China", "Xi'an", 34.25833, 108.92861, "CHD"),
}

CONTAINED_OVERRIDES = [
    (lambda n: "chinese university of hong kong" in n and "shenzhen" in n,
     ("The Chinese University of Hong Kong, Shenzhen", "CN", "China", "Shenzhen", 22.54554, 114.0683, "CUHK-Shenzhen")),
    (lambda n: "chinese university of hong kong" in n,
     ("The Chinese University of Hong Kong", "HK", "Hong Kong", "Hong Kong", 22.27832, 114.17469, "CUHK")),
    (lambda n: "university of electronic science and technology of china" in n,
     ("University of Electronic Science and Technology of China", "CN", "China", "Chengdu", 30.66667, 104.06667, "UESTC")),
    (lambda n: "northwestern polytechnical university" in n,
     ("Northwestern Polytechnical University", "CN", "China", "Xi'an", 34.25833, 108.92861, "NWPU")),
    (lambda n: "city university of hong kong" in n and "shenzhen" not in n and "dongguan" not in n,
     ("City University of Hong Kong", "HK", "Hong Kong", "Hong Kong", 22.27832, 114.17469, "CityUHK")),
    (lambda n: "hong kong polytechnic university" in n,
     ("Hong Kong Polytechnic University", "HK", "Hong Kong", "Hong Kong", 22.27832, 114.17469, "PolyU")),
    (lambda n: ("the university of hong kong" in n or n.startswith("university of hong kong")) and "shenzhen" not in n,
     ("University of Hong Kong", "HK", "Hong Kong", "Pok Fu Lam", 22.26861, 114.12924, "HKU")),
    (lambda n: "university of macau" in n or "university of macao" in n,
     ("University of Macau", "MO", "Macao", "Macao", 22.20056, 113.54611, "UM")),
]

ENTITY_TERMS = {"academy", "college", "company", "corporation", "electronics", "hospital", "institute", "laboratory", "semiconductor", "university"}


def text(value) -> str:
    if value is None:
        return ""
    result = str(value).strip()
    return "" if result.lower() in EMPTY_VALUES else result


def normalize(value: str) -> str:
    value = html.unescape(text(value)).replace("&", " and ")
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch)).lower()
    value = re.sub(r"\b(univ|universite|universitat)\b", "university", value)
    value = re.sub(r"\b(inst|institut)\b", "institute", value)
    value = re.sub(r"\b(technol)\b", "technology", value)
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def query_variants(name: str) -> tuple[list[str], str]:
    base = normalize(name)
    variants = [base]
    country_hint = ""
    for suffix, code in COUNTRY_SUFFIXES.items():
        suffix_norm = normalize(suffix)
        if base.endswith(" " + suffix_norm):
            stripped = base[: -(len(suffix_norm) + 1)].strip()
            if len(stripped) >= 3:
                variants.append(stripped)
                country_hint = code
                break
    return list(dict.fromkeys(variants)), country_hint


def split_ror_names(value: str) -> list[str]:
    result = []
    for item in text(value).split(";"):
        item = re.sub(r"^[a-z]{2,3}(?:_[A-Za-z_]+)?:\s*", "", item.strip())
        if item:
            result.append(item)
    return result


def split_sample_aliases(value: str) -> list[str]:
    result = []
    for item in text(value).split(" | "):
        item = html.unescape(item.strip())
        item = re.sub(r"\s+\(\d+\)\s*$", "", item).strip()
        if item:
            result.append(item)
    return result[:5]


def acronym_for(name: str) -> str:
    words = [w for w in re.findall(r"[A-Za-z0-9]+", name) if w.lower() not in {"of", "the", "and", "for", "at", "in"}]
    if not words:
        return "ORG"
    if len(words) == 1:
        return words[0].upper()[:12]
    result = "".join(w[0].upper() for w in words)
    return result[:12]


def manual_override_for(normalized_name: str):
    exact = MANUAL_OVERRIDES.get(normalized_name)
    if exact:
        return exact
    for predicate, value in CONTAINED_OVERRIDES:
        if predicate(normalized_name):
            return value
    return None


@dataclass
class RorRecord:
    ror_id: str
    display_name: str
    country_code: str
    country_name: str
    city: str
    latitude: float | None
    longitude: float | None
    acronyms: list[str]
    names: list[str]


def float_or_none(value: str) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def load_ror(path: Path):
    records: list[RorRecord] = []
    exact: dict[str, list[int]] = defaultdict(list)
    token_index: dict[str, set[int]] = defaultdict(set)
    country_names: dict[str, str] = {}
    city_points: dict[tuple[str, str], tuple[float, float]] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            display = text(row.get("names.types.ror_display"))
            names = [display]
            names.extend(split_ror_names(row.get("names.types.label", "")))
            names.extend(split_ror_names(row.get("names.types.alias", "")))
            acronyms = split_ror_names(row.get("names.types.acronym", ""))
            names.extend(acronyms)
            names = list(dict.fromkeys(n for n in names if n))
            code = text(row.get("locations.geonames_details.country_code")).upper()
            country = text(row.get("locations.geonames_details.country_name"))
            city = text(row.get("locations.geonames_details.name"))
            lat = float_or_none(row.get("locations.geonames_details.lat"))
            lon = float_or_none(row.get("locations.geonames_details.lng"))
            rec = RorRecord(text(row.get("id")), display, code, country, city, lat, lon, acronyms, names)
            idx = len(records)
            records.append(rec)
            if code and country:
                country_names[code] = country
            if code and city and lat is not None and lon is not None:
                city_points[(code, normalize(city))] = (lat, lon)
            for ror_name in names:
                norm = normalize(ror_name)
                if len(norm) < 2:
                    continue
                exact[norm].append(idx)
                for token in set(norm.split()) - GENERIC_TOKENS:
                    if len(token) >= 2:
                        token_index[token].add(idx)
    city_names: dict[str, set[tuple[str, str]]] = defaultdict(set)
    city_token_index: dict[str, set[str]] = defaultdict(set)
    for code, city_norm in city_points:
        if len(city_norm) >= 4:
            city_names[city_norm].add((code, city_norm))
            for token in city_norm.split():
                if len(token) >= 4:
                    city_token_index[token].add(city_norm)
    return records, exact, token_index, country_names, city_points, city_names, city_token_index


def token_similarity(a: str, b: str) -> float:
    aa, bb = set(a.split()), set(b.split())
    if not aa or not bb:
        return 0.0
    return (2 * len(aa & bb)) / (len(aa) + len(bb))


def name_score(query: str, candidate: str) -> float:
    seq = SequenceMatcher(None, query, candidate).ratio()
    token = token_similarity(query, candidate)
    candidate_meaningful = set(candidate.split()) - GENERIC_TOKENS
    if candidate in query and len(candidate) >= 8 and candidate_meaningful:
        return 0.975
    containment = min(len(query), len(candidate)) / max(len(query), len(candidate)) if query in candidate or candidate in query else 0
    return 0.50 * token + 0.40 * seq + 0.10 * containment


def choose_exact(ids: list[int], records: list[RorRecord], country_hint: str) -> int | None:
    if not ids:
        return None
    if country_hint:
        same = [i for i in ids if records[i].country_code == country_hint]
        if same:
            return same[0]
    if len(ids) == 1:
        return ids[0]
    locations = {(records[i].country_code, normalize(records[i].city)) for i in ids}
    return ids[0] if len(locations) == 1 else None


def infer_country_and_city(values: list[str], country_hint: str, country_names, city_points, city_names, city_token_index):
    joined = " ".join(normalize(v) for v in values if v)
    code = country_hint.upper()
    if not code:
        for phrase, candidate_code in sorted(COUNTRY_SUFFIXES.items(), key=lambda x: -len(x[0])):
            if re.search(rf"\b{re.escape(normalize(phrase))}\b", joined):
                code = candidate_code
                break
    candidate_city_names = set()
    for token in set(joined.split()):
        candidate_city_names.update(city_token_index.get(token, set()))
    candidates = []
    for city_norm in candidate_city_names:
        locations = city_names[city_norm]
        if len(city_norm) < 4 or not re.search(rf"\b{re.escape(city_norm)}\b", joined):
            continue
        for candidate_code, _ in locations:
            if not code or candidate_code == code:
                candidates.append((len(city_norm), candidate_code, city_norm))
    if not candidates:
        return code, country_names.get(code, ""), "", None
    _, candidate_code, city_norm = max(candidates)
    code = code or candidate_code
    coords = city_points.get((code, city_norm))
    display_city = city_norm.title()
    return code, country_names.get(code, ""), display_city, coords


def match_ror(name: str, country_hint: str, records, exact, token_index):
    variants, suffix_hint = query_variants(name)
    country_hint = country_hint or suffix_hint
    for variant in variants:
        chosen = choose_exact(exact.get(variant, []), records, country_hint)
        if chosen is not None:
            return records[chosen], 1.0, "ror-exact"

    query = min(variants, key=len)
    meaningful = [t for t in query.split() if t not in GENERIC_TOKENS and len(t) >= 3]
    token_postings = sorted(
        ((t, token_index.get(t, set())) for t in set(meaningful)),
        key=lambda item: len(item[1]),
    )
    candidates: set[int] = set()
    for _, posting in token_postings[:4]:
        candidates.update(posting)
        if len(candidates) > 2000:
            break
    if not candidates:
        return None, 0.0, "unmatched"

    # Cheap token/country ranking keeps expensive string comparisons bounded.
    query_tokens = set(query.split())
    ranked = []
    for idx in candidates:
        rec = records[idx]
        best_overlap = max(len(query_tokens & set(normalize(n).split())) for n in rec.names)
        country_bonus = 1 if country_hint and rec.country_code == country_hint else 0
        ranked.append((best_overlap, country_bonus, idx))
    candidates = {item[2] for item in sorted(ranked, reverse=True)[:80]}

    best = None
    second = 0.0
    for idx in candidates:
        rec = records[idx]
        score = max(name_score(query, normalize(n)) for n in rec.names)
        if country_hint:
            score += 0.06 if rec.country_code == country_hint else -0.18
        if best is None or score > best[0]:
            second = best[0] if best else second
            best = (score, rec)
        elif score > second:
            second = score
    if best is None:
        return None, 0.0, "unmatched"
    score, rec = best
    margin = score - second
    if score >= 0.86 and margin >= 0.035:
        return rec, min(score, 0.99), "ror-fuzzy-high"
    if score >= 0.80 and margin >= 0.07 and country_hint:
        return rec, min(score, 0.96), "ror-fuzzy-country"
    return None, score, "ambiguous" if score >= 0.72 else "unmatched"


def apply_style(sheet):
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 34
    for i, cell in enumerate(sheet[1], 1):
        width = 15
        if cell.value in {"canonical_name", "cleaned_canonical_name", "sample_raw_aliases", "notes", "review_reason"}:
            width = 42
        sheet.column_dimensions[get_column_letter(i)].width = width


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--ror", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()

    records, exact, token_index, country_names, city_points, city_names, city_token_index = load_ror(args.ror)
    workbook = openpyxl.load_workbook(args.input)
    sheet = workbook["institutions_to_geocode"]
    headers = [text(c.value) for c in sheet[1]]
    required_new = ["original_canonical_name", "cleaned_canonical_name", "ror_id", "matched_ror_name", "match_status", "match_score", "review_reason"]
    for header in required_new:
        if header not in headers:
            headers.append(header)
            sheet.cell(1, len(headers), header)
    col = {name: i + 1 for i, name in enumerate(headers)}

    stats = Counter()
    clean_groups = {}
    review_rows = []
    for row_idx in range(2, sheet.max_row + 1):
        get = lambda name: text(sheet.cell(row_idx, col[name]).value) if name in col else ""
        original = get("canonical_name")
        country_hint = get("resolved_country_code") or get("current_country_code")
        raw_aliases = split_sample_aliases(get("sample_raw_aliases"))
        query_names = [original, *raw_aliases]
        normalized = normalize(original)
        override = manual_override_for(normalized)
        rec = None
        score = 0.0
        status = "unmatched"
        if override:
            display, code, country, city, lat, lon, acronym = override
            status, score = "manual-main-campus", 1.0
            values = (display, code, country, city, lat, lon, acronym, "", display)
        else:
            status_rank = {"ror-exact": 3, "ror-fuzzy-high": 2, "ror-fuzzy-country": 1}
            matches = [match_ror(original, country_hint, records, exact, token_index)]
            if matches[0][2] != "ror-exact":
                for query in raw_aliases:
                    candidate_match = match_ror(query, country_hint, records, exact, token_index)
                    matches.append(candidate_match)
                    if candidate_match[2] == "ror-exact":
                        break
            valid_matches = [m for m in matches if m[0] is not None]
            if valid_matches:
                rec, score, status = max(valid_matches, key=lambda m: (status_rank.get(m[2], 0), m[1]))
            else:
                rec, score, status = max(matches, key=lambda m: m[1])
            if rec:
                acronym = rec.acronyms[0] if rec.acronyms else acronym_for(rec.display_name)
                values = (rec.display_name, rec.country_code, rec.country_name, rec.city, rec.latitude, rec.longitude, acronym, rec.ror_id, rec.display_name)
            else:
                code = country_hint.upper()
                country = get("resolved_country_name") or get("current_country_name") or country_names.get(code, "")
                city = get("resolved_city") or get("current_city")
                inferred_code, inferred_country, inferred_city, inferred_coords = infer_country_and_city(
                    query_names, code, country_names, city_points, city_names, city_token_index
                )
                code = code or inferred_code
                country = country or inferred_country
                city = city or inferred_city
                coords = city_points.get((code, normalize(city))) if code and city else None
                coords = coords or inferred_coords
                lat = float(get("latitude")) if get("latitude") else (coords[0] if coords else None)
                lon = float(get("longitude")) if get("longitude") else (coords[1] if coords else None)
                values = (original, code, country, city, lat, lon, get("学校缩写") or acronym_for(original), "", "")

        cleaned, code, country, city, lat, lon, acronym, ror_id, matched_name = values
        complete_geo = bool(code and country and city and lat is not None and lon is not None)
        if rec or override:
            confidence = round(score * 100)
            source = "ROR v2.9 (2026-06-23)" if rec else "manual-main-campus-rule"
            reason = ""
        else:
            confidence = int(float(get("geo_confidence"))) if get("geo_confidence") else (55 if complete_geo else 25 if code else 0)
            source = get("geo_source") or ("existing-country-city" if complete_geo else "existing-country-only" if code else "")
            reason = "No unique high-confidence ROR match; manual review required"

        assignments = {
            "original_canonical_name": original,
            "cleaned_canonical_name": cleaned,
            "resolved_country_code": code,
            "resolved_country_name": country,
            "resolved_city": city,
            "latitude": lat if lat is not None else "",
            "longitude": lon if lon is not None else "",
            "geo_confidence": confidence,
            "geo_source": source,
            "经纬度": f"{lat:.6f}, {lon:.6f}" if lat is not None and lon is not None else "",
            "学校缩写": acronym,
            "ror_id": ror_id,
            "matched_ror_name": matched_name,
            "match_status": status,
            "match_score": round(score, 4),
            "review_reason": reason,
        }
        for name, value in assignments.items():
            if name in col:
                sheet.cell(row_idx, col[name], value)

        entity_like = bool(set(normalize(cleaned).split()) & ENTITY_TERMS)
        map_ready = complete_geo and bool(ror_id or override or entity_like)
        stats[status] += 1
        stats["map_ready" if map_ready else "excluded_from_clean_master"] += 1
        stats["complete_geo" if complete_geo else "incomplete_geo"] += 1
        stats["has_country" if code else "missing_country"] += 1
        stats["has_city" if city else "missing_city"] += 1
        stats["has_coordinates" if lat is not None and lon is not None else "missing_coordinates"] += 1

        paper_count = int(float(get("paper_count") or 0))
        raw_mentions = int(float(get("raw_mention_count") or 0))
        if map_ready:
            group_key = f"name:{normalize(cleaned)}|{code}|{normalize(city)}"
            if group_key not in clean_groups:
                clean_groups[group_key] = {
                    "institution_id": ror_id or get("normalized_key"), "canonical_name": cleaned,
                    "acronym": acronym, "country_code": code, "country_name": country, "city": city,
                    "latitude": lat, "longitude": lon, "ror_id": ror_id, "paper_count": 0,
                    "raw_mention_count": 0, "source_rows": 0, "geo_confidence": confidence,
                    "match_status": status,
                }
            group = clean_groups[group_key]
            if not group["ror_id"] and ror_id:
                group["ror_id"] = ror_id
                group["institution_id"] = ror_id
            group["paper_count"] += paper_count
            group["raw_mention_count"] += raw_mentions
            group["source_rows"] += 1
            group["geo_confidence"] = max(group["geo_confidence"], confidence)
        if reason:
            review_rows.append([row_idx, original, paper_count, code, city, round(score, 4), reason])

    for name in ["clean_institutions", "needs_review", "quality_summary"]:
        if name in workbook.sheetnames:
            del workbook[name]

    clean_sheet = workbook.create_sheet("clean_institutions")
    clean_headers = ["institution_id", "canonical_name", "acronym", "country_code", "country_name", "city", "latitude", "longitude", "ror_id", "paper_count", "raw_mention_count", "source_rows", "geo_confidence", "match_status"]
    clean_sheet.append(clean_headers)
    for group in sorted(clean_groups.values(), key=lambda x: (-x["paper_count"], x["canonical_name"])):
        clean_sheet.append([group[h] for h in clean_headers])

    review_sheet = workbook.create_sheet("needs_review")
    review_sheet.append(["source_row", "canonical_name", "paper_count", "country_hint", "city_hint", "best_match_score", "review_reason"])
    for item in sorted(review_rows, key=lambda x: (-x[2], x[1])):
        review_sheet.append(item)

    summary_sheet = workbook.create_sheet("quality_summary")
    summary_sheet.append(["metric", "count", "rate"])
    total = sheet.max_row - 1
    for key in sorted(stats):
        summary_sheet.append([key, stats[key], stats[key] / total if total else 0])
    summary_sheet.append(["source_rows", total, 1])
    summary_sheet.append(["deduplicated_institutions", len(clean_groups), len(clean_groups) / total if total else 0])
    summary_sheet.append(["ror_release", "v2.9-2026-06-23", ""])

    for target in [sheet, clean_sheet, review_sheet, summary_sheet]:
        apply_style(target)
    for row in summary_sheet.iter_rows(min_row=2, min_col=3, max_col=3):
        row[0].number_format = "0.00%"

    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)
    print(f"Saved: {args.output}")
    print(f"Source rows: {total}; deduplicated institutions: {len(clean_groups)}")
    for key in sorted(stats):
        print(f"{key}: {stats[key]}")


if __name__ == "__main__":
    main()
