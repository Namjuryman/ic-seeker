#!/usr/bin/env python3
"""Build a strict, parent-level institution master from the enriched workbook."""

from __future__ import annotations

import argparse
import csv
import re
import unicodedata
from collections import Counter, OrderedDict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


INTERNAL_UNIT_PATTERNS = [
    re.compile(r"^state key laborator(?:y|ies)\b", re.I),
    re.compile(r"^(?:school|department|faculty|college|division) of\b", re.I),
    re.compile(r"\b(?:school|department|faculty|college|division) of .+\b(?:university|institute)\b", re.I),
    re.compile(r"^ibm research\s*-", re.I),
    re.compile(r"^stanford systemx alliance$", re.I),
    re.compile(r"^center for advancing electronics dresden$", re.I),
    re.compile(r"^ucl biomedical research centre$", re.I),
    re.compile(r"\b(?:lab|laboratoire)\s*$", re.I),
]

PROTECTED_DISTINCT = {
    "the chinese university of hong kong shenzhen",
    "chinese university of hong kong shenzhen",
}

GENERIC_NAMES = {"faculty united kingdom"}

REQUIRED = [
    "institution_id", "canonical_name", "acronym", "country_code", "country_name",
    "city", "latitude", "longitude", "paper_count", "geo_confidence", "match_status",
]


def norm(value: object) -> str:
    value = unicodedata.normalize("NFKD", str(value or ""))
    value = "".join(ch for ch in value if not unicodedata.combining(ch)).lower()
    return re.sub(r"[^a-z0-9]+", " ", value).strip()


def split_values(value: str) -> list[str]:
    result = []
    for item in str(value or "").split(";"):
        item = re.sub(r"^[a-z]{2,3}(?:_[A-Za-z_]+)?:\s*", "", item.strip())
        if item:
            result.append(item)
    return result


def acronym(name: str, provided: list[str]) -> str:
    if provided:
        return provided[0][:20]
    words = [w for w in re.findall(r"[A-Za-z0-9]+", name) if w.lower() not in {"of", "the", "and", "for", "at", "in"}]
    if not words:
        return "ORG"
    if len(words) == 1:
        return words[0].upper()[:20]
    return "".join(word[0].upper() for word in words)[:20]


def parse_parents(value: str) -> list[str]:
    match = re.search(r"(?:^|; )parent: ([^;]+)", value or "")
    return [part.strip() for part in match.group(1).split(",")] if match else []


def load_ror(path: Path):
    records = {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            records[row["id"]] = {
                "ror_id": row["id"],
                "name": row["names.types.ror_display"],
                "acronym": acronym(row["names.types.ror_display"], split_values(row.get("names.types.acronym", ""))),
                "country_code": row["locations.geonames_details.country_code"],
                "country_name": row["locations.geonames_details.country_name"],
                "city": row["locations.geonames_details.name"],
                "latitude": float(row["locations.geonames_details.lat"]) if row["locations.geonames_details.lat"] else None,
                "longitude": float(row["locations.geonames_details.lng"]) if row["locations.geonames_details.lng"] else None,
                "types": set(split_values(row.get("types", ""))),
                "parents": parse_parents(row.get("relationships", "")),
            }
    return records


def is_internal_unit(name: str) -> bool:
    if norm(name) in PROTECTED_DISTINCT:
        return False
    return any(pattern.search(name) for pattern in INTERNAL_UNIT_PATTERNS)


def choose_parent(record: dict, ror: dict) -> dict | None:
    candidates = [ror[parent_id] for parent_id in record["parents"] if parent_id in ror]
    if not candidates:
        return None
    preferred = [item for item in candidates if item["types"] & {"education", "company"}]
    return (preferred or candidates)[0]


def roll_to_parent(record: dict, ror: dict) -> tuple[dict, list[str]]:
    current = record
    merged = []
    seen = set()
    while current["ror_id"] not in seen and is_internal_unit(current["name"]):
        seen.add(current["ror_id"])
        parent = choose_parent(current, ror)
        if not parent:
            break
        merged.append(current["name"])
        current = parent
    return current, merged


def style_sheet(sheet, wide_columns=()):
    fill = PatternFill("solid", fgColor="17365D")
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 34
    for index, cell in enumerate(sheet[1], 1):
        sheet.column_dimensions[get_column_letter(index)].width = 46 if cell.value in wide_columns else 16


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--ror", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()

    ror = load_ror(args.ror)
    source_wb = openpyxl.load_workbook(args.input, read_only=True, data_only=True)
    source = source_wb["clean_institutions"]
    headers = [cell.value for cell in next(source.iter_rows(min_row=1, max_row=1))]
    index = {name: idx for idx, name in enumerate(headers)}

    groups = OrderedDict()
    merged_units = []
    exclusions = Counter()
    source_count = 0

    for row in source.iter_rows(min_row=2, values_only=True):
        source_count += 1
        item = {name: row[idx] for name, idx in index.items()}
        paper_count = int(item.get("paper_count") or 0)
        trusted = bool(item.get("ror_id")) or item.get("match_status") == "manual-main-campus"
        if paper_count < 1:
            exclusions["zero linked papers"] += 1
            continue
        if not trusted:
            exclusions["no verified institution identity"] += 1
            continue

        canonical = {
            "institution_id": item["institution_id"],
            "canonical_name": item["canonical_name"],
            "acronym": item["acronym"],
            "country_code": item["country_code"],
            "country_name": item["country_name"],
            "city": item["city"],
            "latitude": item["latitude"],
            "longitude": item["longitude"],
            "ror_id": item["ror_id"],
            "paper_count": paper_count,
            "raw_mention_count": int(item.get("raw_mention_count") or 0),
            "source_rows": int(item.get("source_rows") or 0),
            "geo_confidence": item["geo_confidence"],
            "match_status": item["match_status"],
            "merged_subunit_count": 0,
            "merged_subunits": "",
        }

        subunits = []
        if canonical["ror_id"] in ror:
            parent, subunits = roll_to_parent(ror[canonical["ror_id"]], ror)
            if subunits:
                canonical.update({
                    "institution_id": parent["ror_id"],
                    "canonical_name": parent["name"],
                    "acronym": parent["acronym"],
                    "country_code": parent["country_code"],
                    "country_name": parent["country_name"],
                    "city": parent["city"],
                    "latitude": parent["latitude"],
                    "longitude": parent["longitude"],
                    "ror_id": parent["ror_id"],
                    "match_status": "merged-to-parent",
                })
                for subunit in subunits:
                    merged_units.append([subunit, parent["name"], parent["ror_id"], paper_count])

        if is_internal_unit(canonical["canonical_name"]):
            exclusions["unresolved internal subunit"] += 1
            continue
        if norm(canonical["canonical_name"]) in GENERIC_NAMES:
            exclusions["generic organization label"] += 1
            continue

        key = canonical["ror_id"] or f"manual:{norm(canonical['canonical_name'])}|{canonical['country_code']}|{norm(canonical['city'])}"
        if key not in groups:
            groups[key] = canonical
        else:
            group = groups[key]
            group["paper_count"] += canonical["paper_count"]
            group["raw_mention_count"] += canonical["raw_mention_count"]
            group["source_rows"] += canonical["source_rows"]
            group["geo_confidence"] = max(group["geo_confidence"] or 0, canonical["geo_confidence"] or 0)
        if subunits:
            group = groups[key]
            existing = [v for v in group["merged_subunits"].split(" | ") if v]
            group["merged_subunits"] = " | ".join(dict.fromkeys([*existing, *subunits]))
            group["merged_subunit_count"] = len(group["merged_subunits"].split(" | "))

    output_headers = [
        "institution_id", "canonical_name", "acronym", "country_code", "country_name", "city",
        "latitude", "longitude", "ror_id", "paper_count", "raw_mention_count", "source_rows",
        "geo_confidence", "match_status", "merged_subunit_count", "merged_subunits",
    ]
    output_wb = openpyxl.Workbook()
    master = output_wb.active
    master.title = "institution_master"
    master.append(output_headers)
    final_rows = sorted(groups.values(), key=lambda row: (-row["paper_count"], row["canonical_name"]))
    for item in final_rows:
        master.append([item[name] for name in output_headers])

    merged_sheet = output_wb.create_sheet("merged_subunits")
    merged_sheet.append(["subunit_name", "merged_into", "parent_ror_id", "paper_count"])
    for row in sorted(merged_units, key=lambda item: (-item[3], item[0])):
        merged_sheet.append(row)

    summary = output_wb.create_sheet("quality_summary")
    summary.append(["metric", "count", "description"])
    summary_rows = [
        ("source_clean_rows", source_count, "Rows in the prior permissive clean sheet"),
        ("excluded_zero_papers", exclusions["zero linked papers"], "No paper in the 2000-2026 database linked to this row"),
        ("excluded_unverified_identity", exclusions["no verified institution identity"], "No ROR identity or explicit main-campus rule"),
        ("excluded_internal_subunits", exclusions["unresolved internal subunit"], "Internal lab/department with no reliable parent relationship"),
        ("excluded_generic_labels", exclusions["generic organization label"], "Generic label that is not a usable institution identity"),
        ("source_rows_after_strict_filter", source_count - sum(exclusions.values()), "Verified institutions with at least one linked paper"),
        ("subunit_rows_merged", len(merged_units), "Internal labs/departments/centers merged to a parent"),
        ("final_unique_institutions", len(final_rows), "Final institution-level records"),
    ]
    for row in summary_rows:
        summary.append(row)

    readme = output_wb.create_sheet("README")
    readme.append(["Rule", "Meaning"])
    notes = [
        ("Grain", "One row is one verified university, company, research institute, government research body, or healthcare institution."),
        ("Identity requirement", "A row must have a ROR ID or an explicit reviewed main-campus rule."),
        ("Evidence requirement", "A row must link to at least one paper in the 2000-2026 SiliconScope database."),
        ("Internal units", "Departments, schools, State Key Laboratories, and selected corporate/internal centers are merged into their parent."),
        ("Independent institutes", "Formal research institutes such as imec, ETRI, and CAS institutes remain separate institutions."),
        ("Campus exceptions", "CUHK and CUHK-Shenzhen remain distinct; UESTC research-office variants map to the Chengdu main institution."),
        ("Excluded data", "Broken HTML fragments, single-character rows, address fragments, unverified aliases, and zero-paper rows are omitted."),
        ("Geography source", "ROR v2.9 (2026-06-23), with reviewed main-campus exceptions."),
    ]
    for row in notes:
        readme.append(row)

    style_sheet(master, {"canonical_name", "merged_subunits"})
    style_sheet(merged_sheet, {"subunit_name", "merged_into"})
    style_sheet(summary, {"description"})
    style_sheet(readme, {"Meaning"})
    master.sheet_view.showGridLines = False
    args.output.parent.mkdir(parents=True, exist_ok=True)
    output_wb.save(args.output)
    print(f"Saved: {args.output}")
    print(f"Source rows: {source_count}; strict candidates: {source_count - sum(exclusions.values())}; final institutions: {len(final_rows)}")
    print(f"Merged internal units: {len(merged_units)}; exclusions: {dict(exclusions)}")


if __name__ == "__main__":
    main()
